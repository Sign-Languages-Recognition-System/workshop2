from PyQt5.QtWidgets import *
from PyQt5 import uic
import tkinter as tk
from tkinter import messagebox
import sys
from keras.models import load_model
from keras.preprocessing import image
import numpy as np
import cv2
import time
import os
from keras.preprocessing import image
from keras.preprocessing.image import ImageDataGenerator
import tensorflow as tf
import openpyxl 
import sys
import pymsgbox


image_x, image_y = 200,200
correct=0
background = None
accumulated_weight = 0.5
ROI_top = 102
ROI_bottom = 298
ROI_right = 427
ROI_left = 623
user=0
classno=0

word_dict1 = {0:'A',1:'B',2:'C',3:'D',4:'E',5:'F',6:'G',7:'H',8:'I',9:'J',10:'K',11:'L',12:'M',13:'N',14:'O',15:'P',16:'Q',17:'R',18:'S',19:'T',20:'U',21:'V',22:'W',23:'X',24:'Y',25:'Z'}
word_dict2 = {0:'Zero',1:'One',2:'Two',3:'Three',4:'Four',5:'Five',6:'Six',7:'Seven',8:'Eight',9:'Nine'}

model1=load_model('wong_model_test_alpha.h5')
model2=load_model('best_model_overfit_number.h5')


class UI(QMainWindow):
        def __init__(self):
                super(UI,self).__init__()

                uic.loadUi("Main.ui",self)

                self.button1=self.findChild(QPushButton,"Student")
                self.button1.clicked.connect(self.clickedSLogin)

                self.button2=self.findChild(QPushButton,"Teacher")
                self.button2.clicked.connect(self.clickedTLogin)

                self.show()

        def clickedSLogin(self):
            print("Student")
            uic.loadUi("StudentLogin.ui",self)
            self.buttonHome=self.findChild(QPushButton,"back")
            self.buttonHome.clicked.connect(self.clickedHome1)
            self.buttonSubmit=self.findChild(QPushButton,"submit")
            self.buttonSubmit.clicked.connect(self.clickedStudentLogin)
            self.username=self.findChild(QLineEdit,"Sname")
            self.password=self.findChild(QLineEdit,"Spassword")
            self.classname=self.findChild(QLineEdit,"Sclass")

        def clickedTLogin(self):
            print("Teacher")
            uic.loadUi("TeacherLogin.ui",self)
            self.buttonHome=self.findChild(QPushButton,"back2")
            self.buttonHome.clicked.connect(self.clickedHome1)
            self.buttonSubmit=self.findChild(QPushButton,"submit")
            self.buttonSubmit.clicked.connect(self.clickedTeacherLogin)
            self.username=self.findChild(QLineEdit,"Sname")
            self.password=self.findChild(QLineEdit,"Spassword")


        def clickedTeacherLogin(self):
            print("chk1")
            path = "Teacher.xlsx"
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active 
            searchResult=0
            username=self.username.text()
            password=self.password.text()
            for r in range(2,sheet_obj.max_row+1):
                print("chk2")
                searchUsername=sheet_obj.cell(row=r,column=2).value
                searchPassword=sheet_obj.cell(row=r,column=3).value
                if((searchUsername==username) and (searchPassword==password)):
                    found=sheet_obj.cell(row=r,column=1).value
                    print(found)
                    global user
                    user=found+1
                    print("_+_+_+_+_+_+_+_")
                    print(user)
                    searchResult=1
                    break

            if (searchResult==1):
                print("chk3")
                self.clickedBtn2()
            else:
                print("chk4")
                root = tk.Tk()
                root.withdraw()
                messagebox.showerror("LoginError","Incorrect Password/Teacher Name")

            
        def clickedStudentLogin(self):
            print("chk1")
            classname=self.classname.text()
            path = "Student{}.xlsx".format(classname)
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active 
            searchResult=0
            username=self.username.text()
            password=self.password.text()
            for r in range(2,sheet_obj.max_row+1):
                print("chk2")
                searchUsername=sheet_obj.cell(row=r,column=2).value
                searchPassword=sheet_obj.cell(row=r,column=5).value
                if((searchUsername==username) and (searchPassword==password)):
                    found=sheet_obj.cell(row=r,column=1).value
                    print(found)
                    global user
                    user=found+1
                    print("_+_+_+_+_+_+_+_")
                    print(user)
                    searchResult=1
                    global classno
                    classno=classname
                    break

            if (searchResult==1):
                print("chk3")
                self.clickedBtn1()
            else:
                print("chk4")
                root = tk.Tk()
                root.withdraw()
                messagebox.showerror("LoginError","Incorrect Password/Student Name/Class")

        def clickedHome1(self):
            uic.loadUi("Main.ui",self)
            self.button1=self.findChild(QPushButton,"Student")
            self.button1.clicked.connect(self.clickedSLogin)
            self.button2=self.findChild(QPushButton,"Teacher")
            self.button2.clicked.connect(self.clickedTLogin)
            self.show()

        def clickedBtn1(self):
                print("Student")
                uic.loadUi("StudentDash.ui",self)
                self.buttonHome=self.findChild(QPushButton,"HomeButton1")
                self.buttonHome.clicked.connect(self.clickedHome)
                self.button3=self.findChild(QPushButton,"Sbutton2")
                self.button3.clicked.connect(self.clickedTestAlpha)
                self.button4=self.findChild(QPushButton,"Sbutton4")
                self.button4.clicked.connect(self.clickedTestNum)

        def clickedBtn2(self):
                print("Teacher")
                uic.loadUi("TeacherDash.ui",self)
                self.buttonHome=self.findChild(QPushButton,"HomeButton2")
                self.buttonHome.clicked.connect(self.clickedHome)
                self.button5=self.findChild(QPushButton,"Tbutton2")
                self.button5.clicked.connect(self.clickedAlpha)
                self.button6=self.findChild(QPushButton,"Tbutton1")
                self.button6.clicked.connect(self.clickedNum)
                self.button7=self.findChild(QPushButton,"Tbutton3")
                self.button7.clicked.connect(self.clickedReport)
                self.button8=self.findChild(QPushButton,"resetpass")
                self.button8.clicked.connect(self.clickedReset)

        '''def clickedReset(self):
            global user
            print(user)
            os.system("start EXCEL.EXE Teacher.xlsx")'''
                    
        def clickedReset(self):
            print("--------------------------------------------------------")
            uic.loadUi("ResetPassword.ui",self)
            self.buttonHome=self.findChild(QPushButton,"back3")
            self.buttonHome.clicked.connect(self.clickedBtn2)
            self.buttonSubmit=self.findChild(QPushButton,"reset")
            self.buttonSubmit.clicked.connect(self.clickedResetPassword)
            self.password=self.findChild(QLineEdit,"Spassword")
            self.confirmpassword=self.findChild(QLineEdit,"Sconfirmpassword")

        def clickedResetPassword(self):
            global user
            password=self.password.text()
            Confirmpassword=self.confirmpassword.text()
            if(password==Confirmpassword):
                print(user)
                path = "Teacher.xlsx"
                wb_obj = openpyxl.load_workbook(path)
                sheet_obj = wb_obj.active
                sheet_obj.cell(row=user,column=3).value=password
                wb_obj.save("Teacher.xlsx")
                root = tk.Tk()
                root.withdraw()
                messagebox.showinfo("Information","Password changed")
                self.clickedBtn2()
            else:
                root = tk.Tk()
                root.withdraw()
                messagebox.showerror("Error","Password doesn't match")

        def clickedReport(self):
            print("Report")
            '''uic.loadUi("StudentReport.ui",self)
            self.buttonBack=self.findChild(QPushButton,"back")
            self.buttonBack.clicked.connect(self.clickedBtn2)
            self.buttonLoad=self.findChild(QPushButton,"Load")
            self.buttonLoad.clicked.connect(self.clickedLoad)'''
            response = pymsgbox.prompt('Which class report do you wish to see?')
            os.system("start EXCEL.EXE Student{}.xlsx".format(response))

        def clickedTestAlpha(self):
                tpe='a'
                Student(tpe)

        def clickedAlpha(self):
                TeacherAlpha()

        def clickedTestNum(self):
                tpe='n'
                Student(tpe)

        def clickedNum(self):
                TeacherNum()

        def clickedHome(self):
                root = tk.Tk()
                root.overrideredirect(1)
                root.withdraw()
                ans=messagebox.askquestion("Logout", "Are you sure you want to logout?")
                if(ans=='yes'):
                        global user
                        user=0
                        classno=0
                        uic.loadUi("Main.ui",self)

                        self.button1=self.findChild(QPushButton,"Student")
                        self.button1.clicked.connect(self.clickedSLogin)

                        self.button2=self.findChild(QPushButton,"Teacher")
                        self.button2.clicked.connect(self.clickedTLogin)

                        self.show()
                else:
                        root.destroy()


def create_folder(folder_name):
    print("4")
    d = os.path.dirname(__file__) # directory of script
    p = r'{}/'.format(d) # path to be created
    print(p)
    if not os.path.exists(p + folder_name):
        os.mkdir(p + folder_name)

def capture_images(ges_name,alpha,tpe1):
    tpe=tpe1
    print("3")
    global correct
    create_folder(str(ges_name))
    d = os.path.dirname(__file__) # directory of script
    p = r'{}/'.format(d) # path to be created
    print(p)
    cam = cv2.VideoCapture(0)
    cv2.namedWindow("test")
    question=r'Please Show {} and press C button'.format(alpha)
    #img_counter = number
    test_set_image_name = 1
    listImage = [1]
    for loop in listImage:
        while True:
            ret, frame = cam.read()
            frame = cv2.flip(frame, 1)
            if ret:
                   img = cv2.rectangle(frame, (425, 100), (625, 300), (0, 255, 0), thickness=2, lineType=8, shift=0)
                   lower_blue = np.array([0, 0, 0])
                   upper_blue = np.array([179, 25, 255])
                   imcrop = img[102:298, 427:623]
                   hsv = cv2.cvtColor(imcrop, cv2.COLOR_BGR2GRAY)
                   #mask = cv2.inRange(hsv, lower_blue, upper_blue)
                   hsv = cv2.GaussianBlur(hsv, (9, 9), 0)
                   thresh, blackAndWhiteImage = cv2.threshold(hsv, 127, 255, cv2.THRESH_BINARY)
                   #blackAndWhitImage = cv2.bitwise_or(blackAndWhiteImage)
                   #erode = cv2.erode(thresh, kernel, iterations = 1)
                   #result = cv2.bitwise_or(img, erode)
                   #result = cv2.bitwise_or(imcrop, imcrop, mask=mask)
                   cv2.putText(frame, str(question), (30, 400), cv2.FONT_HERSHEY_TRIPLEX, 0.7, (127, 127, 255))
                   #cv2.imshow("testmask", mask)
                   cv2.imshow("test", frame)
                   cv2.imshow("mask", blackAndWhiteImage)
                   #cv2.imshow("result", result)
                   if cv2.waitKey(1) == ord('c'):
                       img_name = p + str(ges_name) + "/1.jpg"
                       save_img = cv2.resize(blackAndWhiteImage, (image_x, image_y))
                       cv2.imwrite(img_name, save_img)
                       print("{} written!".format(img_name))
                       #test_image = image.load_img(p + str(ges_name) +'/1.jpg', target_size=(64, 64))
                       image_folder=p + str(ges_name) +'/1.jpg'
                       frame = cv2.imread(image_folder)
                       lower_black = np.array([0,0,0], dtype = "uint16")
                       upper_black = np.array([70,70,70], dtype = "uint16")
                       black_mask = cv2.inRange(frame, lower_black, upper_black)
                       cv2.imwrite(img_name, black_mask)
                       test_image = image.load_img(p + str(ges_name) +'/1.jpg', target_size=(200, 200))                       
                       test_image = image.img_to_array(test_image)
                       test_image = np.expand_dims(test_image, axis = 0)
                       if(tpe=="a"):
                               result=model1.predict(test_image)
                               answer=word_dict1[np.argmax(result)]
                               print(alpha)
                               print(answer)
                               if(answer==alpha):
                                       correct=correct+1
                                       print(correct)
                       elif(tpe=="n"):
                               result=model2.predict(test_image)
                               answer=word_dict2[np.argmax(result)]
                               print(alpha)
                               print(answer)
                               if(answer==alpha):
                                       correct=correct+1
                                       print(correct)
                       cam.release()
                       cv2.destroyAllWindows()
                       break

def resultPage(c,tpe):
       global classno
       print("******")
       global user
       print(user)
       path = "Student{}.xlsx".format(classno)
       print(path)
       wb_obj = openpyxl.load_workbook(path)
       sheet_obj = wb_obj.active
       if(tpe=="a"):
           sheet_obj.cell(row=user,column=7).value=c
           wb_obj.save(path)
       elif(tpe=="n"):
           sheet_obj.cell(row=user,column=6).value=c
           wb_obj.save(path)
       root = tk.Tk()
       root.overrideredirect(1)
       root.withdraw()
       d=r'Your score is {}/3. You can try again.The newest mark will be recorded'.format(c)
       messagebox.showinfo("Result",d)
       root.destroy()

def startPage():
       root = tk.Tk()
       root.overrideredirect(1)
       root.withdraw()
       ans=messagebox.askquestion("Test", "You have 3 questions, Press C button each time when submit.\nDo you want to proceed?")
       if(ans=='yes'):
              return 1
       else:
              root.destroy()

def Student(tpe):
       tpe1=tpe
       if(startPage()== 1):
              print("1")
              global correct
              print("2")
              correct=0
              if(tpe1=='a'):
                      ges_name='Ans'
                      letter='V'
                      capture_images(ges_name,letter,tpe1)
                      letter='L'
                      capture_images(ges_name,letter,tpe1)
                      letter='B'
                      capture_images(ges_name,letter,tpe1)
                      resultPage(correct,tpe1)
              #sys.exit()
              elif(tpe1=='n'):
                      ges_name='Ans'
                      letter='Three'
                      capture_images(ges_name,letter,tpe1)
                      letter='One'
                      capture_images(ges_name,letter,tpe1)
                      letter='Seven'
                      capture_images(ges_name,letter,tpe1)
                      resultPage(correct,tpe1)

def cal_accum_avg(frame, accumulated_weight):

    global background
    
    if background is None:
        background = frame.copy().astype("float")
        return None

    cv2.accumulateWeighted(frame, background, accumulated_weight)

def segment_hand(frame, threshold=25):
    global background
    
    diff = cv2.absdiff(background.astype("uint8"), frame)

    
    _ , thresholded = cv2.threshold(diff, threshold, 255, cv2.THRESH_BINARY)
    
    #Fetching contours in the frame (These contours can be of hand or any other object in foreground) ...
    contours, hierarchy = cv2.findContours(thresholded.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    # If length of contours list = 0, means we didn't get any contours...
    if len(contours) == 0:
        return None
    else:
        # The largest external contour should be the hand 
        hand_segment_max_cont = max(contours, key=cv2.contourArea)
        
        # Returning the hand segment(max contour) and the thresholded image of hand...
        return (thresholded, hand_segment_max_cont)

def TeacherAlpha():
       cam = cv2.VideoCapture(0)
       num_frames =0
       while True:
              ret, frame = cam.read()

           # filpping the frame to prevent inverted image of captured frame...
              frame = cv2.flip(frame, 1)

              frame_copy = frame.copy()

           # ROI from the frame
              roi = frame[ROI_top:ROI_bottom, ROI_right:ROI_left]

              gray_frame = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
              gray_frame = cv2.GaussianBlur(gray_frame, (9, 9), 0)


              if num_frames < 1:
                     cal_accum_avg(gray_frame, accumulated_weight)
                     cv2.putText(frame_copy, "FETCHING BACKGROUND...PLEASE WAIT", (80, 400), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0,0,255), 2)
    
              else: 
               # segmenting the hand region
                     hand = segment_hand(gray_frame)
        

               # Checking if we are able to detect the hand...
                     if hand is not None:
                            thresholded, hand_segment = hand

                   # Drawing contours around hand segment
                            cv2.drawContours(frame_copy, [hand_segment + (ROI_right, ROI_top)], -1, (255, 0, 0),1)
            
                            cv2.imshow("Thesholded Hand Image", thresholded)
            
                            thresholded = cv2.resize(thresholded, (200, 200))
                            thresholded = cv2.cvtColor(thresholded, cv2.COLOR_GRAY2RGB)
                            thresholded = np.reshape(thresholded, (1,thresholded.shape[0],thresholded.shape[1],3))
            
                            pred = model1.predict(thresholded)
                            cv2.putText(frame_copy, word_dict1[np.argmax(pred)], (170, 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,0,255), 2)
             #else:            
           # Draw ROI on frame_copy
              cv2.rectangle(frame_copy, (ROI_left, ROI_top), (ROI_right, ROI_bottom), (255,128,0), 3)

           # incrementing the number of frames for tracking
              num_frames += 1

           # Display the frame with segmented hand
              cv2.putText(frame_copy, "Press Q to quit", (10, 20), cv2.FONT_ITALIC, 0.5, (51,255,51), 1)
              cv2.imshow("Teaching", frame_copy)

              if cv2.waitKey(1) == ord('q'):
                     cam.release()
                     cv2.destroyAllWindows()
                     break

def TeacherNum():
       cam = cv2.VideoCapture(0)
       num_frames =0
       while True:
              ret, frame = cam.read()

           # filpping the frame to prevent inverted image of captured frame...
              frame = cv2.flip(frame, 1)

              frame_copy = frame.copy()

           # ROI from the frame
              roi = frame[ROI_top:ROI_bottom, ROI_right:ROI_left]

              gray_frame = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
              gray_frame = cv2.GaussianBlur(gray_frame, (9, 9), 0)


              if num_frames < 1:
                     cal_accum_avg(gray_frame, accumulated_weight)
                     cv2.putText(frame_copy, "FETCHING BACKGROUND...PLEASE WAIT", (80, 400), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0,0,255), 2)
    
              else: 
               # segmenting the hand region
                     hand = segment_hand(gray_frame)
        

               # Checking if we are able to detect the hand...
                     if hand is not None:
                            thresholded, hand_segment = hand

                   # Drawing contours around hand segment
                            cv2.drawContours(frame_copy, [hand_segment + (ROI_right, ROI_top)], -1, (255, 0, 0),1)
            
                            cv2.imshow("Thesholded Hand Image", thresholded)
            
                            thresholded = cv2.resize(thresholded, (200, 200))
                            thresholded = cv2.cvtColor(thresholded, cv2.COLOR_GRAY2RGB)
                            thresholded = np.reshape(thresholded, (1,thresholded.shape[0],thresholded.shape[1],3))
            
                            pred = model2.predict(thresholded)
                            cv2.putText(frame_copy, word_dict2[np.argmax(pred)], (170, 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,0,255), 2)
             #else:            
           # Draw ROI on frame_copy
              cv2.rectangle(frame_copy, (ROI_left, ROI_top), (ROI_right, ROI_bottom), (255,128,0), 3)

           # incrementing the number of frames for tracking
              num_frames += 1

           # Display the frame with segmented hand
              cv2.putText(frame_copy, "Press Q to quit", (10, 20), cv2.FONT_ITALIC, 0.5, (51,255,51), 1)
              cv2.imshow("Teaching", frame_copy)

              if cv2.waitKey(1) == ord('q'):
                     cam.release()
                     cv2.destroyAllWindows()
                     break



app=QApplication(sys.argv)
UIWindow=UI()
app.exec_()
